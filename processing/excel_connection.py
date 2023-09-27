import pandas as pd
from openpyxl import load_workbook


class Excel:
    """
    The `Excel` class provides methods for updating and managing existing Excel files, including row-level validation
    and deleting rows based on conditions.

    Attributes:
        excel_file (str): The path to the existing Excel file.
        sheet_name (str): The name of the sheet within the Excel file.

    Methods:
        update_excel_with_row_level_validation(self, external_data, columns_to_verify=None):
            Updates an existing Excel file with row-level validation for specified columns using external data.

        delete_rows_from_excel_multiple_conditions(self, conditions):
            Deletes rows from an existing Excel file based on multiple conditions.

    Use Case:
        The `Excel` class simplifies interactions with Excel files in Python, making it a useful tool for projects that
        involve Excel data management, import, and export.

    Example Usage:
        # Create an Excel object
        excel = Excel(excel_file='data.xlsx', sheet_name='Sheet1')

        # Update Excel with row-level validation
        external_data = pd.read_excel('new_data.xlsx')
        excel.update_excel_with_row_level_validation(external_data, columns_to_verify=['Title', 'Country', 'Indicator'])

        # Delete rows based on conditions  = [(1, 'Value1'), (2, 'Value2')]
        excel.delete_rows_from_excel_multiple_conditions(conditions)
    """

    def __init__(self, excel_file, sheet_name):
        """
        Initialize an Excel object for updating an existing Excel file.

        Parameters:
        - excel_file (str): The path to the existing Excel file.
        - sheet_name (str): The name of the sheet within the Excel file.

        Returns:
        - None
        """
        self.excel_file = excel_file
        self.sheet_name = sheet_name

    def insert_into_existing_excel(self, df, existing_excel=None, sheet_name=None, verify_cols=None):
        """
        Inserts data from a given DataFrame into an existing Excel file and specified sheet, while handling duplicates and providing feedback.

        Parameters:
            df (pandas.DataFrame): The DataFrame containing data to be inserted.
            existing_excel (str): The path to the existing Excel file.
            sheet_name (str): The name of the sheet within the Excel file to insert data into.
            verify_cols (list): A list of column names used to identify duplicates.

        Returns:
            None

        The function reads the existing Excel file specified by 'existing_excel' and the sheet specified by 'sheet_name'. It then checks and enforces data type compatibility between the input DataFrame 'df' and the existing sheet. The 'verify_cols' parameter is used to identify duplicate rows based on specific columns.

        The function counts and prints the number of duplicate rows found in 'df' compared to the existing sheet. It filters 'df' to retain only non-duplicate rows.

        The non-duplicate rows are combined with the existing Excel data, and a 'No.' column is added for row numbering, starting from 1.

        The function appends the combined data to the specified sheet in the existing Excel file.

        If the insertion is successful, the function prints a success message along with the number of rows inserted.
        """

        if existing_excel is None:
            existing_excel = self.excel_file
        if sheet_name is None:
            sheet_name = self.sheet_name

        if verify_cols is None:
            verify_cols = ['Title', 'Country', 'Year', 'Month']
        # Read the existing Excel file
        try:
            df_exit = pd.read_excel(existing_excel, sheet_name=sheet_name)
        except FileNotFoundError:
            print(f"The Excel file '{existing_excel}' does not exist.")
            return
        except KeyError:
            print(f"The sheet '{sheet_name}' does not exist in '{existing_excel}'.")
            return

        # Ensure the data types of columns in df match those in df_exit
        for col, dtype in df_exit.dtypes.items():
            df[col] = df[col].astype(dtype)

        # Concatenate df_exit and df_not_duplicates_in_df_exit
        concatenated_df = pd.concat([df_exit, df], ignore_index=True)
        concatenated_df['No.'] = [i for i in range(1, concatenated_df.shape[0] + 1)]

        num_duplicates = concatenated_df.duplicated(subset=verify_cols).sum()
        print(f"Number of duplicate rows: {num_duplicates}")

        concatenated_df.drop_duplicates(subset=verify_cols, inplace=True)
        concatenated_df['No.'] = [i for i in range(1, concatenated_df.shape[0] + 1)]

        # Create an ExcelWriter object to write to the existing Excel file
        with pd.ExcelWriter(existing_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Write the concatenated DataFrame to the specified sheet
            concatenated_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print('Insert successfully!', concatenated_df.shape[0] - df_exit.shape[0], 'rows')

    def delete_rows_from_excel_multiple_conditions(self, conditions):
        """
        Delete rows from an existing Excel file based on multiple conditions.

        Parameters:
        - excel_file (str): The path to the existing Excel file.
        - sheet_name (str): The name of the sheet from which to delete rows.
        - conditions (list of tuples): A list of (condition_column_index, condition_value) tuples.

        Returns:
        - None
        """
        excel_file, sheet_name = self.excel_file, self.sheet_name

        try:
            # Load the existing Excel file
            wb = load_workbook(excel_file)

            # Select the worksheet
            sheet = wb[sheet_name]

            # Store the row indices to delete
            rows_to_delete = []

            conditions = list(conditions)

            # Iterate through rows to identify rows to delete
            for row_index, row in enumerate(sheet.iter_rows(min_row=0, values_only=True)):
                for condition_column, condition_value in conditions:
                    if row[condition_column] == condition_value:  # Adjust column index to 0-based
                        rows_to_delete.append(row_index + 1)

            # Delete rows in reverse order to avoid issues with row indices
            for row_index in reversed(rows_to_delete):
                sheet.delete_rows(row_index)

            # Save the modified Excel file
            wb.save(excel_file)
            wb.close()
            print("Rows deleted successfully.")
        except Exception as e:
            print(f"An error occurred: {str(e)}")

    def show_excel(self):
        return pd.read_excel(self.excel_file, sheet_name=self.sheet_name)
